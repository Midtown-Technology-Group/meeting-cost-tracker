"""
Console reporting for meeting cost analytics.
"""

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import track
from rich import box

from .calculator import MeetingCost, CostAnalytics


class ConsoleReporter:
    """Generate rich console output for analytics."""
    
    def __init__(self, console: Console):
        self.console = console
    
    def display_summary(self, analytics: CostAnalytics):
        """Display summary statistics."""
        self.console.print(Panel(
            f"[bold green]${analytics.total_cost:,.2f}[/bold green] total cost\n"
            f"[cyan]{analytics.total_meetings}[/cyan] meetings\n"
            f"[cyan]{analytics.total_meeting_hours:.1f}[/cyan] hours\n"
            f"[cyan]{analytics.total_attendee_hours:.1f}[/cyan] attendee-hours\n\n"
            f"Average: [yellow]${analytics.average_meeting_cost:.2f}[/yellow] per meeting, "
            f"[yellow]{analytics.average_meeting_duration:.1f} min[/yellow] duration, "
            f"[yellow]{analytics.average_attendees_per_meeting:.1f}[/yellow] attendees",
            title="Meeting Cost Summary",
            border_style="green",
        ))
    
    def display_top_meetings(self, meetings: list[MeetingCost]):
        """Display table of most expensive meetings."""
        if not meetings:
            self.console.print("[dim]No meetings to display.[/dim]")
            return
        
        table = Table(
            title="Most Expensive Meetings",
            box=box.ROUNDED,
            show_header=True,
            header_style="bold",
        )
        
        table.add_column("Rank", style="dim", width=4)
        table.add_column("Date", style="cyan", width=12)
        table.add_column("Subject", style="white", min_width=30)
        table.add_column("Duration", style="blue", width=10)
        table.add_column("Attendees", style="blue", width=9)
        table.add_column("Cost", style="green", justify="right", width=12)
        table.add_column("$/hr", style="dim", justify="right", width=10)
        
        for i, mc in enumerate(meetings, 1):
            date = mc.meeting.start_time.strftime("%Y-%m-%d")
            subject = mc.meeting.subject[:40] + "..." if len(mc.meeting.subject) > 40 else mc.meeting.subject
            duration = f"{mc.meeting.duration_minutes:.0f} min"
            
            # Color code by cost
            cost_color = "green"
            if mc.total_cost > 1000:
                cost_color = "red"
            elif mc.total_cost > 500:
                cost_color = "yellow"
            
            table.add_row(
                str(i),
                date,
                subject,
                duration,
                str(mc.attendee_count),
                f"[bold {cost_color}]${mc.total_cost:,.2f}[/bold {cost_color}]",
                f"${mc.cost_per_hour:,.2f}",
            )
        
        self.console.print(table)
    
    def display_top_attendees(self, attendees: list[tuple]):
        """Display table of top attendees by cost."""
        if not attendees:
            return
        
        table = Table(
            title="Top Attendees by Cost",
            box=box.ROUNDED,
            show_header=True,
            header_style="bold",
        )
        
        table.add_column("Rank", style="dim", width=4)
        table.add_column("Attendee", style="cyan", min_width=30)
        table.add_column("Total Cost", style="green", justify="right", width=12)
        table.add_column("Meetings", style="blue", width=10)
        table.add_column("Avg per Meeting", style="dim", justify="right", width=15)
        
        for i, (email, cost, count) in enumerate(attendees, 1):
            avg = cost / count if count > 0 else 0
            table.add_row(
                str(i),
                email[:35],
                f"${cost:,.2f}",
                str(count),
                f"${avg:,.2f}",
            )
        
        self.console.print(table)
    
    def display_savings(self, analytics: CostAnalytics):
        """Display potential savings analysis."""
        if analytics.potential_savings_30min == 0 and analytics.potential_savings_15min == 0:
            return
        
        self.console.print(Panel(
            f"[bold]Potential Savings:[/bold]\n\n"
            f"If all meetings were [cyan]30 minutes[/cyan] or less:\n"
            f"  💰 Save [bold green]${analytics.potential_savings_30min:,.2f}[/bold green]\n\n"
            f"If all meetings were [cyan]15 minutes shorter[/cyan]:\n"
            f"  💰 Save [bold green]${analytics.potential_savings_15min:,.2f}[/bold green]",
            title="Optimization Opportunities",
            border_style="yellow",
        ))


class ExcelReporter:
    """Generate Excel reports with charts."""
    
    def export(self, analytics: CostAnalytics, meeting_costs: list, path: str):
        """Export analytics to Excel workbook."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.console.print("[red]Error: pandas and openpyxl required for Excel export[/red]")
            return
        
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Cost", f"${analytics.total_cost:,.2f}"],
            ["Total Meetings", analytics.total_meetings],
            ["Total Hours", f"{analytics.total_meeting_hours:.1f}"],
            ["Attendee-Hours", f"{analytics.total_attendee_hours:.1f}"],
            ["Avg Cost per Meeting", f"${analytics.average_meeting_cost:.2f}"],
            ["Avg Duration (min)", f"{analytics.average_meeting_duration * 60:.0f}"],
            ["Avg Attendees", f"{analytics.average_attendees_per_meeting:.1f}"],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Meetings sheet
        ws_meetings = wb.create_sheet("Meetings")
        meetings_data = []
        for mc in meeting_costs:
            meetings_data.append({
                "Date": mc.meeting.start_time.strftime("%Y-%m-%d"),
                "Subject": mc.meeting.subject,
                "Duration (min)": mc.meeting.duration_minutes,
                "Attendees": mc.attendee_count,
                "Total Cost": mc.total_cost,
                "Cost/Hour": mc.cost_per_hour,
                "Organizer": mc.meeting.organizer.email if mc.meeting.organizer else "",
            })
        
        df_meetings = pd.DataFrame(meetings_data)
        for r in dataframe_to_rows(df_meetings, index=False, header=True):
            ws_meetings.append(r)
        
        # Attendees sheet
        ws_attendees = wb.create_sheet("Attendees")
        attendees_data = []
        for email, cost, count in analytics.top_attendees_by_cost:
            attendees_data.append({
                "Attendee": email,
                "Total Cost": cost,
                "Meeting Count": count,
                "Average per Meeting": cost / count if count > 0 else 0,
            })
        
        df_attendees = pd.DataFrame(attendees_data)
        for r in dataframe_to_rows(df_attendees, index=False, header=True):
            ws_attendees.append(r)
        
        # Daily costs sheet
        ws_daily = wb.create_sheet("Daily Costs")
        daily_data = [["Date", "Total Cost"]]
        for date, cost in sorted(analytics.cost_by_day.items()):
            daily_data.append([date, cost])
        
        for row in daily_data:
            ws_daily.append(row)
        
        # Add chart to daily costs
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Meeting Costs by Day"
        chart.y_axis.title = 'Cost ($)'
        chart.x_axis.title = 'Date'
        
        data = Reference(ws_daily, min_col=2, min_row=1, max_row=len(daily_data), max_col=2)
        cats = Reference(ws_daily, min_col=1, min_row=2, max_row=len(daily_data))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_daily.add_chart(chart, "D2")
        
        # Save
        wb.save(path)
